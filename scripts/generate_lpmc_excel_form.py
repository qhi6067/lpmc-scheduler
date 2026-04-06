from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


TITLE = "Las Palmas Medical Center"
DEFAULT_SUBTITLE = "Tech Schedule"
DEFAULT_START_DATE = date(2026, 4, 12)
DEFAULT_END_DATE = date(2026, 5, 24)
DEFAULT_OUTPUT_PATH = Path("samples") / "Las_Palmas_Tech_Schedule_Form.xlsx"
PAYDAY_ANCHOR = date(2026, 4, 24)

TECH_ROWS = [
    "FABIOLA",
    "FRANK",
    "LUIS",
    "",
    "DAPHNE",
    "SARA",
    "ALEXANDRA",
    "ARTURO",
    "",
    "MIRIAM",
    "HILDA",
    "REBECCA",
    "NICOLE",
    "HIMILCE",
    "JASMIN",
    "JULIANA",
    "JOE",
    "",
    "AARON",
    "ZULMALYZ",
    "STEPHANIE",
    "ALYSSA",
    "AILYN",
    "BRANDEN",
    "RICARDO",
    "DAISY",
    "MARIA",
    "JAIME",
    "NIA",
]

PHARMACIST_ROWS = [
    "ADRIAN",
    "BIANCA",
    "CARLOS",
    "DANIELA",
    "ELENA",
    "FERNANDO",
    "",
    "GABRIELA",
    "HECTOR",
    "ISABEL",
    "JULIO",
    "KAREN",
]

SHIFT_CODES = [
    ("M", "0630-1500"),
    ("1", "0630-1500 RUNNER"),
    ("2", "0600-1430 IV"),
    ("3", "1330-2200 IV"),
    ("2C", "0600-1430 CHEMO IV"),
    ("4", "1130-2000"),
    ("4W", "1000-1830"),
    ("5", "1330-2200 RUNNER"),
    ("6", "1900-0630"),
    ("6A", "1800-0630"),
    ("6B", "1800-0630"),
    ("9", "7:30-1600 duties depend on volunteer status"),
    ("A", "AUDITS"),
    ("E", "Expiration/station checks 10-1830"),
    ("P", "0730-1600"),
    ("PO", "Requested off / PTO"),
    ("RH", "Regional / resource hospital"),
    ("ED", "Emergency department"),
    ("RH/ED", "0830-1500"),
    ("MT", "Misc. code shown on source schedule"),
    ("V", "Vacation / unavailable"),
    ("r", "Receiving in ABC/MCK"),
]


def build_dates(start_date: date, end_date: date) -> list[date]:
    dates: list[date] = []
    current = start_date
    while current <= end_date:
        dates.append(current)
        current += timedelta(days=1)
    return dates


def build_pay_dates(start_date: date, end_date: date) -> set[date]:
    pay_dates: set[date] = set()

    delta_days = (start_date - PAYDAY_ANCHOR).days
    steps = delta_days // 14
    current = PAYDAY_ANCHOR + timedelta(days=steps * 14)

    while current < start_date:
        current += timedelta(days=14)

    while current <= end_date:
        pay_dates.add(current)
        current += timedelta(days=14)

    return pay_dates


def format_short_date(value: date) -> str:
    return f"{value.month}/{value.day}/{value.strftime('%y')}"


def apply_grid_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border


def generate_schedule_form(
    *,
    subtitle: str = DEFAULT_SUBTITLE,
    start_date: date = DEFAULT_START_DATE,
    end_date: date = DEFAULT_END_DATE,
    employee_rows: list[str] | tuple[str, ...] = TECH_ROWS,
    output_path: Path = DEFAULT_OUTPUT_PATH,
) -> Path:
    if end_date < start_date:
        raise ValueError("end_date must be on or after start_date")

    dates = build_dates(start_date, end_date)
    pay_dates = build_pay_dates(start_date, end_date)
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "Schedule Form"
    lists = workbook.create_sheet("Lists")
    legend = workbook.create_sheet("Legend")

    last_date_col = 1 + len(dates)
    last_date_letter = get_column_letter(last_date_col)

    schedule.sheet_view.showGridLines = False
    schedule.freeze_panes = "B7"
    schedule.column_dimensions["A"].width = 18
    for column in range(2, last_date_col + 1):
        schedule.column_dimensions[get_column_letter(column)].width = 4.2

    for row in range(1, 80):
        schedule.row_dimensions[row].height = 20

    title_fill = PatternFill("solid", fgColor="D9E2F3")
    header_fill = PatternFill("solid", fgColor="EDEDED")
    group_fill = PatternFill("solid", fgColor="F7F7F7")
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrapped_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    schedule.merge_cells(f"B2:{last_date_letter}2")
    schedule["B2"] = TITLE
    schedule["B2"].font = Font(size=15, bold=True)
    schedule["B2"].alignment = center
    schedule["B2"].fill = title_fill

    schedule.merge_cells(f"B3:{last_date_letter}3")
    schedule["B3"] = f"{subtitle} | {format_short_date(start_date)} - {format_short_date(end_date)}"
    schedule["B3"].font = Font(size=12, bold=True)
    schedule["B3"].alignment = center
    schedule["B3"].fill = title_fill

    schedule["A5"] = "Name/Date"
    schedule["A5"].font = Font(bold=True)
    schedule["A5"].alignment = left
    schedule["A5"].fill = header_fill

    schedule["A6"] = "DATE:"
    schedule["A6"].font = Font(bold=True)
    schedule["A6"].alignment = left
    schedule["A6"].fill = header_fill

    weekday_map = ["M", "T", "W", "TH", "F", "S", "S"]
    for index, day in enumerate(dates, start=2):
        pay_cell = schedule.cell(row=4, column=index)
        if day in pay_dates:
            pay_cell.value = "PAY"
            pay_cell.font = Font(size=8, bold=True)
            pay_cell.alignment = center

        weekday_cell = schedule.cell(row=5, column=index)
        weekday_cell.value = weekday_map[day.weekday()]
        weekday_cell.font = Font(bold=True)
        weekday_cell.alignment = center
        weekday_cell.fill = header_fill

        date_cell = schedule.cell(row=6, column=index)
        date_cell.value = day.day
        date_cell.font = Font(bold=True)
        date_cell.alignment = center
        date_cell.fill = header_fill

    start_row = 7
    blank_rows: list[int] = []
    editable_rows: list[int] = []
    for offset, employee_name in enumerate(employee_rows):
        row_number = start_row + offset
        name_cell = schedule.cell(row=row_number, column=1)
        name_cell.value = employee_name
        name_cell.alignment = left
        if employee_name:
            name_cell.font = Font(bold=True)
            editable_rows.append(row_number)
        else:
            blank_rows.append(row_number)
            for column in range(1, last_date_col + 1):
                schedule.cell(row=row_number, column=column).fill = group_fill

    end_row = start_row + len(employee_rows) - 1

    apply_grid_borders(schedule, 5, end_row, 1, last_date_col)

    for row_number in blank_rows:
        for column in range(1, last_date_col + 1):
            schedule.cell(row=row_number, column=column).fill = group_fill

    lists.sheet_state = "hidden"
    for row_number, (code, description) in enumerate(SHIFT_CODES, start=1):
        lists.cell(row=row_number, column=1, value=code)
        lists.cell(row=row_number, column=2, value=description)

    validation = DataValidation(
        type="list",
        formula1=f"=Lists!$A$1:$A${len(SHIFT_CODES)}",
        allow_blank=True,
    )
    validation.promptTitle = "Shift code"
    validation.prompt = "Pick a code from the dropdown or type a custom note."
    validation.errorTitle = "Invalid code"
    validation.error = "Use one of the listed schedule codes or clear the cell."
    schedule.add_data_validation(validation)

    for row_number in editable_rows:
        validation.add(f"B{row_number}:{last_date_letter}{row_number}")
        for column in range(2, last_date_col + 1):
            schedule.cell(row=row_number, column=column).alignment = center

    note_row = end_row + 3
    schedule.merge_cells(f"A{note_row}:{last_date_letter}{note_row + 1}")
    schedule[f"A{note_row}"] = (
        "Editable Excel form based on the posted schedule. Fill or adjust the day cells as needed. "
        "The parser reads from the Name/Date row downward, so the date range is managed separately."
    )
    schedule[f"A{note_row}"].alignment = wrapped_left
    schedule[f"A{note_row}"].fill = note_fill
    schedule[f"A{note_row}"].font = Font(italic=True)

    responsibility_row = note_row + 3
    schedule.merge_cells(f"A{responsibility_row}:{last_date_letter}{responsibility_row + 1}")
    schedule[f"A{responsibility_row}"] = (
        "It is the responsibility of the technician to cover their shift and find coverage "
        "if they choose to make changes to the schedule. Vacation requests are due by the "
        "middle Wednesday of the schedule."
    )
    schedule[f"A{responsibility_row}"].alignment = wrapped_left
    schedule[f"A{responsibility_row}"].fill = note_fill
    schedule[f"A{responsibility_row}"].font = Font(bold=True)

    for cell_ref in ("B2", "B3", "A5", "A6"):
        schedule[cell_ref].border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

    schedule.print_area = f"A1:{last_date_letter}{responsibility_row + 1}"
    schedule.page_setup.orientation = "landscape"
    schedule.page_setup.fitToWidth = 1
    schedule.page_setup.fitToHeight = 0

    legend.sheet_view.showGridLines = False
    legend.column_dimensions["A"].width = 10
    legend.column_dimensions["B"].width = 42
    legend.merge_cells("A1:B1")
    legend["A1"] = "Shift Code Legend"
    legend["A1"].font = Font(size=14, bold=True)
    legend["A1"].alignment = center
    legend["A1"].fill = title_fill
    legend["A2"] = "Code"
    legend["B2"] = "Meaning"
    legend["A2"].font = Font(bold=True)
    legend["B2"].font = Font(bold=True)
    legend["A2"].fill = header_fill
    legend["B2"].fill = header_fill
    legend["A2"].alignment = center
    legend["B2"].alignment = center

    for row_number, (code, description) in enumerate(SHIFT_CODES, start=3):
        legend.cell(row=row_number, column=1, value=code)
        legend.cell(row=row_number, column=2, value=description)
        legend.cell(row=row_number, column=1).alignment = center
        legend.cell(row=row_number, column=2).alignment = left

    legend_note_row = len(SHIFT_CODES) + 5
    legend.merge_cells(f"A{legend_note_row}:B{legend_note_row + 2}")
    legend[f"A{legend_note_row}"] = (
        "Some short codes such as MT, RH, ED, V, and PO are included because they appear in the "
        "photo, even when the printed legend is partially abbreviated."
    )
    legend[f"A{legend_note_row}"].alignment = wrapped_left
    legend[f"A{legend_note_row}"].fill = note_fill
    legend[f"A{legend_note_row}"].font = Font(italic=True)
    apply_grid_borders(legend, 2, len(SHIFT_CODES) + 2, 1, 2)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path.resolve()


def main() -> None:
    result_path = generate_schedule_form()
    print(result_path)


if __name__ == "__main__":
    main()
